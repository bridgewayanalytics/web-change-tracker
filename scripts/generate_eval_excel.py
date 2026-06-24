"""
Generate an auditable accuracy evaluation Excel file.

For each analyzed alert (excluding legacy-schema rows 11 & 20):
  - Blue header row: row #, alert type, date, run_id, agent_call_id
  - Gray "Agent Said" row: human-readable field values (no raw JSON)
  - White "Score / Correction" row:
      ✓  = Correct (2)
      △  = Partial — only used when genuinely ambiguous
      ✗  = Wrong (0) — used decisively, including for newsreel yes/no errors
      N/A = Field not applicable to this alert type

Run from repo root:
    python3 scripts/generate_eval_excel.py
"""

import json
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("pip install openpyxl")

DATA_PATH = Path("analysis/accuracy_eval/eval_data.json")
OUT_PATH  = Path("analysis/accuracy_eval/accuracy_audit.xlsx")

EXCLUDE_ROWS = {11, 20}   # legacy-schema rows — null fields, wrong types

# ── 14 scored fields ──────────────────────────────────────────────────────────
FIELDS = [
    ("alert_type",                                   "Alert Type"),
    ("alert_title",                                  "Alert Title"),
    ("alert_description",                            "Alert Description"),
    ("organization",                                 "Organization"),
    ("event_title",                                  "Event Title"),
    ("event_start_date_time",                        "Event Start"),
    ("event_end_date_time",                          "Event End"),
    ("event_url",                                    "Event URL"),
    ("event_call_in_number_access_code",             "Call-In / Code"),
    ("library_item_preliminary_title",               "Document Title"),
    ("library_item_url",                             "Document URL"),
    ("library_items_file_name",                      "Document Filename"),
    ("agenda_item_title_chronicle_topics",           "Agenda Items & Chronicle Topics"),
    ("is_the_alert_relevant_for_an_art_newsreel_article", "Newsreel Relevant?"),
]
FIELD_KEYS   = [f[0] for f in FIELDS]
FIELD_LABELS = [f[1] for f in FIELDS]

# ── Human-readable value rendering ────────────────────────────────────────────
def _fmt_datetime(s: str) -> str:
    """2026-07-15T15:00:00-04:00 → July 15, 2026  3:00 PM ET"""
    if not s or s in ("N/A", "null", "None"):
        return s or "N/A"
    try:
        from datetime import datetime, timezone, timedelta
        dt = datetime.fromisoformat(s)
        offset = dt.utcoffset()
        if offset == timedelta(hours=-4):
            tz_label = "ET"
        elif offset == timedelta(hours=-5):
            tz_label = "ET"
        else:
            tz_label = str(dt.tzinfo)
        return dt.strftime(f"%B %-d, %Y  %-I:%M %p {tz_label}")
    except Exception:
        return s

def readable_val(key: str, val) -> str:
    if val is None:
        return "N/A"
    if key in ("event_start_date_time", "event_end_date_time"):
        return _fmt_datetime(str(val))
    if key == "organization":
        if isinstance(val, list):
            return "\n".join(val)
        return str(val)
    if key == "library_item_preliminary_title":
        if isinstance(val, dict):
            title  = val.get("title", "N/A")
            status = val.get("status", "")
            if title == "N/A":
                return "N/A"
            return f"{title}  [{status}]" if status else title
        return str(val)
    if key == "is_the_alert_relevant_for_an_art_newsreel_article":
        if isinstance(val, dict):
            status  = val.get("status", "")
            details = val.get("details", "")
            if details and details not in ("N/A", "null", "false", "true"):
                return f"{status} — {details}"
            return status
        return str(val)
    if key == "agenda_item_title_chronicle_topics":
        if not isinstance(val, list):
            return str(val) if val else "N/A"
        lines = []
        for item in val:
            if not isinstance(item, dict):
                lines.append(str(item))
                continue
            title   = item.get("agenda_item_title", "")
            topics  = item.get("chronicle_topics", [])
            status  = item.get("status", "")
            topic_s = ", ".join(t for t in topics if t) if topics else "N/A"
            prefix  = f"[{status}] " if status else ""
            lines.append(f"{prefix}{title}\n  Topics: {topic_s}")
        return "\n\n".join(lines) if lines else "N/A"
    if isinstance(val, list):
        return "\n".join(str(v) for v in val)
    if isinstance(val, dict):
        return str(val)
    return str(val)


# ── Scores & corrections ──────────────────────────────────────────────────────
# score: 2 = correct  |  1 = partial (genuinely ambiguous only)  |  0 = wrong
#        "N/A" = field not applicable to this alert type
#
# Newsreel scoring is binary: correct yes/no = 2, wrong yes/no = 0.
# Event title: any agent-invented text not on the NAIC page = 0.
# Chronicle topics: N/A when a clear topic exists in the taxonomy = 0.

SCORES = {
    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 1 — New or Updated Report or Other Resource
    # ═══════════════════════════════════════════════════════════════════════════
    (1,"alert_type"):   (2,""),
    (1,"alert_title"):  (2,""),
    (1,"alert_description"): (2,""),
    (1,"organization"): (2,""),
    (1,"event_title"):  ("N/A",""),
    (1,"event_start_date_time"): ("N/A",""),
    (1,"event_end_date_time"):   ("N/A",""),
    (1,"event_url"):    ("N/A",""),
    (1,"event_call_in_number_access_code"): ("N/A",""),
    (1,"library_item_preliminary_title"): (2,""),
    (1,"library_item_url"):  (2,""),
    (1,"library_items_file_name"): (2,""),
    (1,"agenda_item_title_chronicle_topics"): (0,
        "chronicle_topics = N/A — should assign a specific topic from the taxonomy, "
        "e.g. 'Health Insurance' or 'SHIP / MHP Funding'"),
    (1,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 2 — Other (Newsroom article swap)
    # ═══════════════════════════════════════════════════════════════════════════
    (2,"alert_type"):   (2,""),
    (2,"alert_title"):  (2,""),
    (2,"alert_description"): (2,""),
    (2,"organization"): (2,""),
    (2,"event_title"):  ("N/A",""),
    (2,"event_start_date_time"): ("N/A",""),
    (2,"event_end_date_time"):   ("N/A",""),
    (2,"event_url"):    ("N/A",""),
    (2,"event_call_in_number_access_code"): ("N/A",""),
    (2,"library_item_preliminary_title"): ("N/A",""),
    (2,"library_item_url"):  ("N/A",""),
    (2,"library_items_file_name"): ("N/A",""),
    (2,"agenda_item_title_chronicle_topics"): (2,""),
    (2,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 3 — Updated Request for Comment (Health RBC)
    # ═══════════════════════════════════════════════════════════════════════════
    (3,"alert_type"):   (2,""),
    (3,"alert_title"):  (2,""),
    (3,"alert_description"): (2,""),
    (3,"organization"): (2,""),
    (3,"event_title"):  ("N/A",""),
    (3,"event_start_date_time"): ("N/A",""),
    (3,"event_end_date_time"):   ("N/A",""),
    (3,"event_url"):    ("N/A",""),
    (3,"event_call_in_number_access_code"): ("N/A",""),
    (3,"library_item_preliminary_title"): (2,""),
    (3,"library_item_url"):  (2,""),
    (3,"library_items_file_name"): (2,""),
    (3,"agenda_item_title_chronicle_topics"): (0,
        "chronicle_topics = N/A — should tag 'Risk-Based Capital' and/or 'Health RBC'"),
    (3,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 4 — New Meeting (HRBC WG July 15)
    # ═══════════════════════════════════════════════════════════════════════════
    (4,"alert_type"):   (2,""),
    (4,"alert_title"):  (2,""),
    (4,"alert_description"): (2,""),
    (4,"organization"): (2,""),
    (4,"event_title"):  (0,
        "Agent invented the prefix 'NAIC Interim Meeting:' — this text does not appear "
        "on the NAIC page. Page says: 'Health Risk-Based Capital (E) Working Group Public Webex Meeting'"),
    (4,"event_start_date_time"): (2,""),
    (4,"event_end_date_time"):   (2,""),
    (4,"event_url"):    (2,""),
    (4,"event_call_in_number_access_code"): (2,""),
    (4,"library_item_preliminary_title"): ("N/A",""),
    (4,"library_item_url"):  ("N/A",""),
    (4,"library_items_file_name"): ("N/A",""),
    (4,"agenda_item_title_chronicle_topics"): (2,""),
    (4,"is_the_alert_relevant_for_an_art_newsreel_article"): (0,
        "WRONG — A new meeting being scheduled is not a newsreel article. "
        "Newsreels cover decisions made and substantive regulatory developments, not calendar announcements. "
        "Should be: No"),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 5 — New Agenda & Materials (P&C RBC joint meeting)
    # ═══════════════════════════════════════════════════════════════════════════
    (5,"alert_type"):   (2,""),
    (5,"alert_title"):  (2,""),
    (5,"alert_description"): (2,""),
    (5,"organization"): (0,
        "Joint meeting — only one of two organizations captured. Missing: 'Catastrophe Risk (E) Subgroup'. "
        "Should be both: Property & Casualty Risk-Based Capital (E) Working Group AND Catastrophe Risk (E) Subgroup"),
    (5,"event_title"):  (2,""),
    (5,"event_start_date_time"): (2,""),
    (5,"event_end_date_time"):   (2,""),
    (5,"event_url"):    (2,""),
    (5,"event_call_in_number_access_code"): (2,""),
    (5,"library_item_preliminary_title"): (2,""),
    (5,"library_item_url"):  (2,""),
    (5,"library_items_file_name"): (2,""),
    (5,"agenda_item_title_chronicle_topics"): (2,""),
    (5,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 6 — New Request for Comment (GOES Model Change Templates)
    # ═══════════════════════════════════════════════════════════════════════════
    (6,"alert_type"):   (2,""),
    (6,"alert_title"):  (2,""),
    (6,"alert_description"): (1,
        "Too brief to be useful — 'New library item posted as part of the 2026 GOES exposure materials for comment.' "
        "Should describe what the GOES Model Change Templates document contains and what stakeholders are being asked to comment on"),
    (6,"organization"): (2,""),
    (6,"event_title"):  ("N/A",""),
    (6,"event_start_date_time"): ("N/A",""),
    (6,"event_end_date_time"):   ("N/A",""),
    (6,"event_url"):    ("N/A",""),
    (6,"event_call_in_number_access_code"): ("N/A",""),
    (6,"library_item_preliminary_title"): (0,
        "Raw filename used as title: 'GOES Model Change Templates 052826_0.xlsx'. "
        "Should be a human-readable name, e.g. 'GOES Model Change Templates (May 2026 Exposure)'"),
    (6,"library_item_url"):  (2,""),
    (6,"library_items_file_name"): (2,""),
    (6,"agenda_item_title_chronicle_topics"): (2,""),
    (6,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 7 — Updated Meeting (LATF June 25 canceled)
    # ═══════════════════════════════════════════════════════════════════════════
    (7,"alert_type"):   (2,""),
    (7,"alert_title"):  (2,""),
    (7,"alert_description"): (2,""),
    (7,"organization"): (2,""),
    (7,"event_title"):  (0,
        "Agent invented a date-code suffix not on the NAIC page: 'Life Actuarial (A) Task Force - 06252026'. "
        "Should be: 'Life Actuarial (A) Task Force Public Conference Call — June 25, 2026'"),
    (7,"event_start_date_time"): (2,""),
    (7,"event_end_date_time"):   (2,""),
    (7,"event_url"):    (2,""),
    (7,"event_call_in_number_access_code"): (2,""),
    (7,"library_item_preliminary_title"): ("N/A",""),
    (7,"library_item_url"):  ("N/A",""),
    (7,"library_items_file_name"): ("N/A",""),
    (7,"agenda_item_title_chronicle_topics"): (2,""),
    (7,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 8 — New Materials (ACLI comment letter, collateral loans)
    # ═══════════════════════════════════════════════════════════════════════════
    (8,"alert_type"):   (2,""),
    (8,"alert_title"):  (2,""),
    (8,"alert_description"): (2,""),
    (8,"organization"): (2,""),
    (8,"event_title"):  (2,""),
    (8,"event_start_date_time"): (2,""),
    (8,"event_end_date_time"):   (2,""),
    (8,"event_url"):    (2,""),
    (8,"event_call_in_number_access_code"): (2,""),
    (8,"library_item_preliminary_title"): (2,""),
    (8,"library_item_url"):  (2,""),
    (8,"library_items_file_name"): (2,""),
    (8,"agenda_item_title_chronicle_topics"): (2,""),
    (8,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 9 — Updated Materials (AG 55 + meeting shortened)
    # ═══════════════════════════════════════════════════════════════════════════
    (9,"alert_type"):   (2,""),
    (9,"alert_title"):  (2,""),
    (9,"alert_description"): (2,""),
    (9,"organization"): (2,""),
    (9,"event_title"):  (0,
        "Mixed invented format — 'NAIC Interim Meeting: Life Actuarial (A) Task Force - 06182026'. "
        "Contains both an invented prefix AND an invented date-code suffix, neither of which appears on the NAIC page. "
        "Should be: 'Life Actuarial (A) Task Force Public Webex Meeting — June 18, 2026'"),
    (9,"event_start_date_time"): (2,""),
    (9,"event_end_date_time"):   (2,""),
    (9,"event_url"):    (2,""),
    (9,"event_call_in_number_access_code"): (2,""),
    (9,"library_item_preliminary_title"): (2,""),
    (9,"library_item_url"):  (2,""),
    (9,"library_items_file_name"): (2,""),
    (9,"agenda_item_title_chronicle_topics"): (0,
        "chronicle_topics = N/A for AG 55 review — should tag 'Actuarial Guidelines' "
        "and/or 'Principle-Based Reserving (PBR)'"),
    (9,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 10 — Updated Agenda & Materials (LATF June 11, 7 materials)
    # ═══════════════════════════════════════════════════════════════════════════
    (10,"alert_type"):   (2,""),
    (10,"alert_title"):  (2,""),
    (10,"alert_description"): (2,""),
    (10,"organization"): (2,""),
    (10,"event_title"):  (2,""),
    (10,"event_start_date_time"): (2,""),
    (10,"event_end_date_time"):   (2,""),
    (10,"event_url"):    (2,""),
    (10,"event_call_in_number_access_code"): (2,""),
    (10,"library_item_preliminary_title"): (1,
        "Only the first of 7 posted materials is captured: 'APF 2025-14 VA Scope Clarification'. "
        "The other 6 (APF 2026-04, APF 2026-05, APF 2025-18, APF 2025-20, ACLI letter, Group Annuity memo) "
        "are described in the alert_description but not represented in this field. "
        "This is a flat-schema limitation — one document slot per row."),
    (10,"library_item_url"):  (2,""),
    (10,"library_items_file_name"): (2,""),
    (10,"agenda_item_title_chronicle_topics"): (2,""),
    (10,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 12 — New Agenda (Financial Stability TF, S&P Private Credit)
    # ═══════════════════════════════════════════════════════════════════════════
    (12,"alert_type"):   (2,""),
    (12,"alert_title"):  (2,""),
    (12,"alert_description"): (2,""),
    (12,"organization"): (2,""),
    (12,"event_title"):  (0,
        "Invented date-code suffix — 'Financial Stability (E) Task Force - 06152026'. "
        "Should be: 'Financial Stability (E) Task Force Public Webex Meeting — June 15, 2026'"),
    (12,"event_start_date_time"): (2,""),
    (12,"event_end_date_time"):   (2,""),
    (12,"event_url"):    (2,""),
    (12,"event_call_in_number_access_code"): (2,""),
    (12,"library_item_preliminary_title"): (0,
        "Returns N/A — but the alert description says 'a new agenda document was posted'. "
        "The agenda PDF should be captured here as the library item with its URL"),
    (12,"library_item_url"):  (0,
        "Returns N/A — should be the URL of the agenda PDF that was posted"),
    (12,"library_items_file_name"): (0,
        "Returns N/A — should be the filename of the agenda PDF"),
    (12,"agenda_item_title_chronicle_topics"): (0,
        "chronicle_topics = N/A for an S&P presentation on private credit — "
        "should tag 'Capital Markets' and/or 'Private Equity / Private Credit'"),
    (12,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 13 — New or Updated Report or Other Resource (House appropriations letter)
    # ═══════════════════════════════════════════════════════════════════════════
    (13,"alert_type"):   (2,""),
    (13,"alert_title"):  (2,""),
    (13,"alert_description"): (2,""),
    (13,"organization"): (2,""),
    (13,"event_title"):  ("N/A",""),
    (13,"event_start_date_time"): ("N/A",""),
    (13,"event_end_date_time"):   ("N/A",""),
    (13,"event_url"):    ("N/A",""),
    (13,"event_call_in_number_access_code"): ("N/A",""),
    (13,"library_item_preliminary_title"): (2,""),
    (13,"library_item_url"):  (2,""),
    (13,"library_items_file_name"): (2,""),
    (13,"agenda_item_title_chronicle_topics"): (0,
        "Agenda item fields all N/A — for a report/resource alert, the document subject should "
        "be tagged with a chronicle topic, e.g. 'Health Insurance' or 'SHIP / MHP Appropriations'"),
    (13,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 14 — Other (past LRBC meeting entry removed)
    # ═══════════════════════════════════════════════════════════════════════════
    (14,"alert_type"):   (2,""),
    (14,"alert_title"):  (2,""),
    (14,"alert_description"): (2,""),
    (14,"organization"): (2,""),
    (14,"event_title"):  (0,
        "Invented prefix 'Old event:' not on the NAIC page. "
        "Should use standard format or a flag like '[REMOVED] Life Risk-Based Capital (E) Working Group — June 11, 2026'"),
    (14,"event_start_date_time"): (2,""),
    (14,"event_end_date_time"):   (2,""),
    (14,"event_url"):    (2,""),
    (14,"event_call_in_number_access_code"): (2,""),
    (14,"library_item_preliminary_title"): (2,""),
    (14,"library_item_url"):  (2,""),
    (14,"library_items_file_name"): (2,""),
    (14,"agenda_item_title_chronicle_topics"): (0,
        "chronicle_topics = N/A — the removed meeting covered collateral loans proposals; "
        "should tag 'Collateral Loans'"),
    (14,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 15 — Updated Request for Comment (LATF expired exposure drafts)
    # ═══════════════════════════════════════════════════════════════════════════
    (15,"alert_type"):   (2,""),
    (15,"alert_title"):  (2,""),
    (15,"alert_description"): (2,""),
    (15,"organization"): (2,""),
    (15,"event_title"):  ("N/A",""),
    (15,"event_start_date_time"): ("N/A",""),
    (15,"event_end_date_time"):   ("N/A",""),
    (15,"event_url"):    ("N/A",""),
    (15,"event_call_in_number_access_code"): ("N/A",""),
    (15,"library_item_preliminary_title"): (0,
        "Title is just the proposal number: 'APF 2025-14'. "
        "Should be the full name: 'APF 2025-14 — VM Variable Annuity Scope Clarification'"),
    (15,"library_item_url"):  (2,""),
    (15,"library_items_file_name"): (2,""),
    (15,"agenda_item_title_chronicle_topics"): (2,""),
    (15,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 16 — MISCLASSIFIED as New Meeting (was a Newsroom article)
    # ═══════════════════════════════════════════════════════════════════════════
    (16,"alert_type"):   (0,
        "WRONG — The NAIC Newsroom page had a new article appear about the Spring National Meeting. "
        "This is not a new calendar event. Should be: 'Other'"),
    (16,"alert_title"):  (2,""),
    (16,"alert_description"): (2,""),
    (16,"organization"): (0,
        "Full legal name used: 'National Association of Insurance Commissioners'. "
        "Should match org tree: 'NAIC'"),
    (16,"event_title"):  (0,
        "WRONG — 'NAIC Spring National Meeting' was pulled from article text, not from a calendar event posting. "
        "No event exists on the Newsroom page. Should be: N/A"),
    (16,"event_start_date_time"): (0,
        "WRONG context — this alert is a newsroom article, not an event. "
        "Returning N/A for start date is technically correct (no dates found) "
        "but the alert_type misclassification makes this field meaningless"),
    (16,"event_end_date_time"):   (0,
        "Same — wrong classification context; field is meaningless here"),
    (16,"event_url"):    ("N/A",""),
    (16,"event_call_in_number_access_code"): ("N/A",""),
    (16,"library_item_preliminary_title"): ("N/A",""),
    (16,"library_item_url"):  ("N/A",""),
    (16,"library_items_file_name"): ("N/A",""),
    (16,"agenda_item_title_chronicle_topics"): (0,
        "chronicle_topics = N/A for a national meeting announcement — "
        "could tag relevant regulatory areas covered at the Spring National Meeting"),
    (16,"is_the_alert_relevant_for_an_art_newsreel_article"): (0,
        "WRONG — A general national meeting preview article is not a newsreel. "
        "Newsreels cover specific regulatory decisions and developments. "
        "Should be: No"),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 17 — New Agenda & Materials (Reinsurance TF, negative IMR)
    # ═══════════════════════════════════════════════════════════════════════════
    (17,"alert_type"):   (2,""),
    (17,"alert_title"):  (2,""),
    (17,"alert_description"): (2,""),
    (17,"organization"): (2,""),
    (17,"event_title"):  (2,""),
    (17,"event_start_date_time"): (2,""),
    (17,"event_end_date_time"):   (2,""),
    (17,"event_url"):    (2,""),
    (17,"event_call_in_number_access_code"): (2,""),
    (17,"library_item_preliminary_title"): (2,""),
    (17,"library_item_url"):  (2,""),
    (17,"library_items_file_name"): (2,""),
    (17,"agenda_item_title_chronicle_topics"): (2,""),
    (17,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 18 — New Request for Comment (GOES, First 10 Path Comparison)
    # ═══════════════════════════════════════════════════════════════════════════
    (18,"alert_type"):   (2,""),
    (18,"alert_title"):  (2,""),
    (18,"alert_description"): (1,
        "Too brief — same sparse description as Row 6: 'New library item posted as part of the 2026 GOES exposure materials for comment.' "
        "Should describe what 'First 10 Path Comparison' contains and what stakeholders are asked to review"),
    (18,"organization"): (2,""),
    (18,"event_title"):  ("N/A",""),
    (18,"event_start_date_time"): ("N/A",""),
    (18,"event_end_date_time"):   ("N/A",""),
    (18,"event_url"):    ("N/A",""),
    (18,"event_call_in_number_access_code"): ("N/A",""),
    (18,"library_item_preliminary_title"): (2,""),
    (18,"library_item_url"):  (2,""),
    (18,"library_items_file_name"): (2,""),
    (18,"agenda_item_title_chronicle_topics"): (2,""),
    (18,"is_the_alert_relevant_for_an_art_newsreel_article"): (2,""),

    # ═══════════════════════════════════════════════════════════════════════════
    # ROW 19 — Updated Meeting (LATF June 18 extended to 1.5 hours)
    # ═══════════════════════════════════════════════════════════════════════════
    (19,"alert_type"):   (2,""),
    (19,"alert_title"):  (2,""),
    (19,"alert_description"): (2,""),
    (19,"organization"): (2,""),
    (19,"event_title"):  (0,
        "Invented date-code suffix — 'NAIC Interim Meeting: Life Actuarial (A) Task Force - 06182026'. "
        "Should be: 'Life Actuarial (A) Task Force Public Webex Meeting — June 18, 2026'"),
    (19,"event_start_date_time"): (2,""),
    (19,"event_end_date_time"):   (2,""),
    (19,"event_url"):    (2,""),
    (19,"event_call_in_number_access_code"): (2,""),
    (19,"library_item_preliminary_title"): ("N/A",""),
    (19,"library_item_url"):  ("N/A",""),
    (19,"library_items_file_name"): ("N/A",""),
    (19,"agenda_item_title_chronicle_topics"): (2,""),
    (19,"is_the_alert_relevant_for_an_art_newsreel_article"): (0,
        "WRONG — A 30-minute meeting extension is not a newsreel article. "
        "Newsreels cover substantive regulatory decisions, not scheduling changes. "
        "Should be: No"),
}

# ── Styles ────────────────────────────────────────────────────────────────────
GREEN      = PatternFill("solid", fgColor="C6EFCE")
YELLOW     = PatternFill("solid", fgColor="FFEB9C")
RED        = PatternFill("solid", fgColor="FFC7CE")
GRAY_NA    = PatternFill("solid", fgColor="E0E0E0")
HEADER_BG  = PatternFill("solid", fgColor="1F3864")
AGENT_BG   = PatternFill("solid", fgColor="F5F5F5")
ROW_HDR_BG = PatternFill("solid", fgColor="2F5496")

thin   = Side(style="thin",   color="BBBBBB")
thick  = Side(style="medium", color="888888")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
THICK_TOP = Border(left=thin, right=thin, top=thick, bottom=thin)

def score_fill(score):
    if score == 2:     return GREEN
    if score == 1:     return YELLOW
    if score == 0:     return RED
    if score == "N/A": return GRAY_NA
    return PatternFill("solid", fgColor="FFFFFF")

def set_cell(ws, row, col, value="", fill=None, bold=False, wrap=True,
             size=9, color="000000", border=BORDER, italic=False):
    c = ws.cell(row=row, column=col, value=value)
    if fill:
        c.fill = fill
    c.font = Font(bold=bold, italic=italic, size=size, color=color)
    c.alignment = Alignment(wrap_text=wrap, vertical="top")
    c.border = border
    return c

# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    with open(DATA_PATH) as f:
        data = json.load(f)

    wb = openpyxl.Workbook()

    # ── Sheet 1: Audit ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Accuracy Audit"

    # Column layout:
    #   A  = row label ("Agent Said" / "Score / Correction")
    #   B  = Run ID  (informational, not scored)
    #   C+ = 14 scored fields
    RUN_ID_COL = 2
    col_offset  = 3          # scored fields start at column C
    total_cols  = col_offset + len(FIELDS) - 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22   # Run ID
    widths = [16, 32, 48, 30, 32, 18, 18, 42, 16, 32, 48, 36, 58, 36]
    for i, w in enumerate(widths):
        ws.column_dimensions[get_column_letter(col_offset + i)].width = w

    # Column headers (row 1)
    set_cell(ws, 1, 1, "", fill=HEADER_BG, bold=True, size=9, color="FFFFFF")
    set_cell(ws, 1, RUN_ID_COL, "Run ID", fill=HEADER_BG, bold=True, size=9, color="FFFFFF")
    for i, label in enumerate(FIELD_LABELS):
        set_cell(ws, 1, col_offset + i, label, fill=HEADER_BG, bold=True, size=9, color="FFFFFF")
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "C2"

    current_row = 2

    for rec in data:
        idx = rec["index"]
        if idx in EXCLUDE_ROWS:
            continue

        alert  = rec["alert"]
        atype  = alert.get("alert_type", "")
        adate  = str(alert.get("alert_date_time", ""))[:10]
        run_id = alert.get("run_id", "")

        # ── Row group header (single clean line) ──────────────────────────────
        header_val = f"ROW {idx}  ·  {atype}  ·  {adate}"
        hc = ws.cell(row=current_row, column=1, value=header_val)
        hc.fill = ROW_HDR_BG
        hc.font = Font(bold=True, size=9, color="FFFFFF")
        hc.alignment = Alignment(wrap_text=False, vertical="center")
        hc.border = THICK_TOP
        ws.merge_cells(
            start_row=current_row, start_column=1,
            end_row=current_row, end_column=total_cols
        )
        ws.row_dimensions[current_row].height = 20
        current_row += 1

        # ── "Agent Said" row ──────────────────────────────────────────────────
        agent_row = current_row
        set_cell(ws, agent_row, 1, "Agent Said", fill=AGENT_BG, bold=True, size=8.5)
        set_cell(ws, agent_row, RUN_ID_COL, run_id, fill=AGENT_BG, size=8)
        for i, key in enumerate(FIELD_KEYS):
            val = readable_val(key, alert.get(key))
            set_cell(ws, agent_row, col_offset + i, val, fill=AGENT_BG, size=8.5)
        ws.row_dimensions[agent_row].height = 72
        current_row += 1

        # ── Score / Correction row ────────────────────────────────────────────
        score_row = current_row
        set_cell(ws, score_row, 1, "Score / Correction", fill=None, bold=True,
                 italic=True, size=8.5, color="555555")
        # Run ID cell in score row — blank (not scored)
        set_cell(ws, score_row, RUN_ID_COL, "", fill=GRAY_NA, size=8.5)
        for i, key in enumerate(FIELD_KEYS):
            entry = SCORES.get((idx, key), (2, ""))
            score, note = entry

            if score == "N/A":
                cell_text = "N/A"
            elif score == 2:
                cell_text = "✓  Correct"
            elif score == 1:
                cell_text = f"△  Partial\n{note}"
            elif score == 0:
                cell_text = f"✗  Wrong\n{note}"
            else:
                cell_text = str(score)

            set_cell(ws, score_row, col_offset + i, cell_text,
                     fill=score_fill(score), size=8.5,
                     bold=(score == 0))
        ws.row_dimensions[score_row].height = 72
        current_row += 1

        # ── Thin spacer ───────────────────────────────────────────────────────
        ws.row_dimensions[current_row].height = 5
        current_row += 1

    # ── Sheet 2: Field Accuracy Summary ───────────────────────────────────────
    ws2 = wb.create_sheet("Field Accuracy Summary")
    ws2.column_dimensions["A"].width = 46
    ws2.column_dimensions["B"].width = 14
    ws2.column_dimensions["C"].width = 14
    ws2.column_dimensions["D"].width = 12

    hdr_labels = ["Field", "Score / Max", "Rows Scored", "Accuracy"]
    for i, h in enumerate(hdr_labels):
        c = ws2.cell(row=1, column=i+1, value=h)
        c.fill = HEADER_BG
        c.font = Font(bold=True, color="FFFFFF", size=10)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER
    ws2.row_dimensions[1].height = 24

    field_totals = {}
    for key in FIELD_KEYS:
        scored = [(s, n) for (ridx, fk), (s, n) in SCORES.items()
                  if fk == key and ridx not in EXCLUDE_ROWS and s != "N/A"]
        if scored:
            total = sum(s for s, _ in scored)
            maxx  = len(scored) * 2
            field_totals[key] = (total, maxx, len(scored))
        else:
            field_totals[key] = (0, 0, 0)

    # Sort by accuracy descending
    rows_data = []
    for key, label in FIELDS:
        total, maxx, n = field_totals[key]
        pct = round(total / maxx * 100, 1) if maxx else 0.0
        rows_data.append((label, total, maxx, n, pct))
    rows_data.sort(key=lambda x: -x[4])

    for r, (label, total, maxx, n, pct) in enumerate(rows_data, 2):
        ws2.cell(row=r, column=1, value=label).font = Font(size=10)
        ws2.cell(row=r, column=2, value=f"{total} / {maxx}").alignment = Alignment(horizontal="center")
        ws2.cell(row=r, column=3, value=n).alignment = Alignment(horizontal="center")
        pc = ws2.cell(row=r, column=4, value=f"{pct}%")
        pc.alignment = Alignment(horizontal="center")
        pc.font = Font(bold=True, size=10)
        if pct >= 95:
            pc.fill = GREEN
        elif pct >= 80:
            pc.fill = PatternFill("solid", fgColor="D9EAD3")
        elif pct >= 60:
            pc.fill = YELLOW
        else:
            pc.fill = RED
        for col in range(1, 5):
            ws2.cell(row=r, column=col).border = BORDER
        ws2.row_dimensions[r].height = 20

    # Overall row
    all_totals = [(s, n) for (ridx, fk), (s, n) in SCORES.items()
                  if ridx not in EXCLUDE_ROWS and s != "N/A"]
    grand_total = sum(s for s, _ in all_totals)
    grand_max   = len(all_totals) * 2
    grand_pct   = round(grand_total / grand_max * 100, 1)
    sep_row = len(rows_data) + 2
    ws2.row_dimensions[sep_row].height = 8
    ov_row = sep_row + 1
    ws2.cell(row=ov_row, column=1, value="OVERALL  (18 current-schema rows)").font = Font(bold=True, size=11)
    ws2.cell(row=ov_row, column=2, value=f"{grand_total} / {grand_max}").font = Font(bold=True)
    ws2.cell(row=ov_row, column=2).alignment = Alignment(horizontal="center")
    ws2.cell(row=ov_row, column=3, value=18).alignment = Alignment(horizontal="center")
    oc = ws2.cell(row=ov_row, column=4, value=f"{grand_pct}%")
    oc.font = Font(bold=True, size=12)
    oc.alignment = Alignment(horizontal="center")
    oc.fill = GREEN if grand_pct >= 90 else YELLOW
    ws2.row_dimensions[ov_row].height = 26

    # Note on excluded rows
    note_row = ov_row + 2
    ws2.cell(row=note_row, column=1,
             value="Rows 11 & 20 excluded — generated with legacy schema (null fields, wrong types). Need rerun.").font = Font(italic=True, color="888888", size=9)

    OUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_PATH)
    print(f"Saved → {OUT_PATH}")
    print(f"Overall accuracy (18 rows): {grand_pct}%  ({grand_total} / {grand_max})")
    print("\nPer-field breakdown:")
    for label, total, maxx, n, pct in rows_data:
        print(f"  {label:<48} {pct:5.1f}%  ({total}/{maxx}, {n} rows)")

if __name__ == "__main__":
    main()
