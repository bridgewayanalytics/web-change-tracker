"""
Generate a scored eval Excel workbook for document extraction agent accuracy.

Evaluates the newsreel_relevance field specifically, since that's the only
field currently in document_extractions_table.jsonl that maps to a Yes/No
correctness judgment without requiring the full chronicle taxonomy.

Usage
-----
    python3 scripts/generate_eval_excel_doc_extraction.py
"""

import json
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    raise SystemExit("pip install openpyxl")

DATA_PATH   = Path("analysis/accuracy_eval/eval_data_doc_extraction.json")
OUTPUT_PATH = Path("analysis/accuracy_eval/accuracy_audit_doc_extraction.xlsx")

# ─── Pre-scored judgements ────────────────────────────────────────────────────
# Tuple: (ground_truth_verdict, score_0_to_2, notes)
#   score 2 = correct
#   score 1 = borderline / arguable either way
#   score 0 = wrong
#
# Ground truth reasoning:
#   ART newsreel covers actuarial & insurance regulatory developments.
#   Yes = document warrants a newsreel article; No = routine/operational doc.
#
NR_SCORES = {
    # index: (ground_truth, score, notes)
    1:  ("Yes", 2, "Capital Markets Bureau investment report — substantive analysis, clearly newsworthy"),
    2:  ("No",  2, "APP Manual routine update — reference document, not a news event"),
    3:  ("Yes", 2, "Web article about Spring National Meeting — announcement/recap is newsreel material"),
    4:  ("No",  0, "International calendar of scheduling dates — operational doc, not newsreel content"),
    5:  ("Yes", 0, "NAIC letter to Congress on SHIP/MHP grants — policy advocacy is newsworthy; agent said No"),
    6:  ("No",  0, "Blanks WG routine meeting agenda — operational doc; agent said Yes incorrectly"),
    7:  ("Yes", 2, "Gov Affairs brief on GCC/LST/Holding Company Act — substantive regulatory update"),
    8:  ("Yes", 0, "NAIC letter on AHP/stop-loss/telehealth bills — NAIC positions on legislation = newsworthy; agent said No"),
    9:  ("No",  0, "VM editorial corrections — non-substantive technical edit; agent said Yes incorrectly"),
    10: ("Yes", 2, "APF 2025-14 VA scope clarification — substantive actuarial proposal"),
}

# ─── Style helpers ────────────────────────────────────────────────────────────
GREEN  = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFEB9C")
RED    = PatternFill("solid", fgColor="FFC7CE")
BLUE   = PatternFill("solid", fgColor="DDEEFF")
HEADER = PatternFill("solid", fgColor="2F5496")
SUBHD  = PatternFill("solid", fgColor="9DC3E6")
WHITE  = PatternFill("solid", fgColor="FFFFFF")

HDR_FONT  = Font(bold=True, color="FFFFFF", size=11)
BOLD      = Font(bold=True, size=10)
NORM      = Font(size=10)
SMALL     = Font(size=9)

THIN = Border(
    left=Side(style="thin"),   right=Side(style="thin"),
    top=Side(style="thin"),    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="top", wrap_text=True)
WRAP   = Alignment(horizontal="left",   vertical="top", wrap_text=True)
WRAP_C = Alignment(horizontal="center", vertical="top", wrap_text=True)


def hdr(ws, row, col, val, width=None):
    c = ws.cell(row=row, column=col, value=val)
    c.font = HDR_FONT; c.fill = HEADER; c.alignment = WRAP_C; c.border = THIN
    if width:
        ws.column_dimensions[get_column_letter(col)].width = width


def subhdr(ws, row, col, val):
    c = ws.cell(row=row, column=col, value=val)
    c.font = BOLD; c.fill = SUBHD; c.alignment = WRAP_C; c.border = THIN


def cell(ws, row, col, val, fill=None, font=None, align=None):
    c = ws.cell(row=row, column=col, value=val)
    c.fill  = fill  or WHITE
    c.font  = font  or NORM
    c.alignment = align or WRAP
    c.border = THIN
    return c


# ─── Sheet 1: Newsreel Relevance Scoring ─────────────────────────────────────
def sheet_newsreel(wb, records):
    ws = wb.create_sheet("Newsreel Relevance")

    # Header row
    headers = [
        ("#", 4), ("Title", 38), ("Doc Type", 16), ("Org", 20),
        ("Agent Says", 11), ("Ground Truth", 13), ("Score", 8), ("Notes", 40),
    ]
    for col, (h, w) in enumerate(headers, 1):
        hdr(ws, 1, col, h, w)
    ws.row_dimensions[1].height = 22

    total_score = 0
    max_score   = 0

    for r in records:
        row_out = r["index"] + 1
        e       = r["extraction"]
        agent_nr = r["newsreel_status"]

        gt, score, notes = NR_SCORES.get(r["index"], ("?", 0, ""))
        total_score += score
        max_score   += 2

        fill = GREEN if score == 2 else (YELLOW if score == 1 else RED)

        title = str(e.get("document_title") or e.get("library_item_title") or "")[:80]
        doc_type = str(e.get("document_type") or "")
        org_val  = e.get("organization_or_publisher")
        org_str  = org_val.get("name") if isinstance(org_val, dict) else str(org_val or "")

        cell(ws, row_out, 1, r["index"],   fill, align=WRAP_C)
        cell(ws, row_out, 2, title,        fill)
        cell(ws, row_out, 3, doc_type,     fill)
        cell(ws, row_out, 4, org_str,      fill)
        cell(ws, row_out, 5, agent_nr,     fill, align=WRAP_C)
        cell(ws, row_out, 6, gt,           fill, align=WRAP_C)
        cell(ws, row_out, 7, f"{score}/2", fill, align=WRAP_C)
        cell(ws, row_out, 8, notes,        fill)
        ws.row_dimensions[row_out].height = 42

    # Summary row
    summary_row = len(records) + 2
    pct = round(100 * total_score / max_score, 1) if max_score else 0
    ws.merge_cells(start_row=summary_row, start_column=1,
                   end_row=summary_row, end_column=6)
    c = ws.cell(row=summary_row, column=1,
                value=f"TOTAL: {total_score}/{max_score} = {pct}% accuracy")
    c.font = Font(bold=True, size=11)
    c.fill = SUBHD; c.alignment = CENTER; c.border = THIN
    return pct


# ─── Sheet 2: Document Detail ─────────────────────────────────────────────────
def sheet_detail(wb, records):
    ws = wb.create_sheet("Document Detail")

    headers = [
        ("#", 4), ("Title", 40), ("Type", 18), ("Description", 50),
        ("Org", 22), ("Newsreel (Agent)", 16), ("Details", 35),
        ("Doc URL", 30), ("Page Length", 12),
    ]
    for col, (h, w) in enumerate(headers, 1):
        hdr(ws, 1, col, h, w)

    for r in records:
        row_out = r["index"] + 1
        e       = r["extraction"]
        _, score, _ = NR_SCORES.get(r["index"], ("?", 0, ""))
        fill = GREEN if score == 2 else (YELLOW if score == 1 else RED)

        org_val = e.get("organization_or_publisher")
        org_str = org_val.get("name") if isinstance(org_val, dict) else str(org_val or "")

        title = str(e.get("document_title") or e.get("library_item_title") or "")
        desc  = str(e.get("document_description") or "")[:300]
        nr_val = e.get("newsreel_relevance")
        nr_details = r.get("newsreel_details") or (
            nr_val.get("details") if isinstance(nr_val, dict) else ""
        ) or ""

        cell(ws, row_out, 1, r["index"],          fill, align=WRAP_C)
        cell(ws, row_out, 2, title,               fill)
        cell(ws, row_out, 3, str(e.get("document_type") or ""), fill)
        cell(ws, row_out, 4, desc,                fill)
        cell(ws, row_out, 5, org_str,             fill)
        cell(ws, row_out, 6, r["newsreel_status"], fill, align=WRAP_C)
        cell(ws, row_out, 7, str(nr_details)[:200], fill)
        cell(ws, row_out, 8, r["doc_url"][:80],   fill)
        cell(ws, row_out, 9, r["doc_page_length"], fill, align=WRAP_C)
        ws.row_dimensions[row_out].height = 60

    ws.freeze_panes = "A2"


# ─── Sheet 3: Summary ────────────────────────────────────────────────────────
def sheet_summary(wb, records, pct):
    ws = wb.create_sheet("Summary", 0)
    ws.sheet_view.showGridLines = False

    def big(row, col, val, fill=None, font=None):
        c = ws.cell(row=row, column=col, value=val)
        c.fill  = fill  or WHITE
        c.font  = font  or NORM
        c.alignment = WRAP
        c.border = THIN
        return c

    ws.merge_cells("A1:F1")
    c = ws.cell(row=1, column=1, value="Document Extraction Agent — Newsreel Relevance Accuracy")
    c.font = Font(bold=True, size=14, color="2F5496"); c.alignment = CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:F2")
    c = ws.cell(row=2, column=1, value="2026-06-24 | 10 doc extraction rows | Targeted: balanced Yes/No sample")
    c.font = Font(size=10, italic=True, color="595959"); c.alignment = CENTER
    ws.row_dimensions[2].height = 18

    # Score box
    ws.merge_cells("A4:B5")
    score_fill = GREEN if pct >= 80 else (YELLOW if pct >= 65 else RED)
    c = ws.cell(row=4, column=1, value=f"{pct}%")
    c.font = Font(bold=True, size=28); c.fill = score_fill; c.alignment = CENTER; c.border = THIN

    ws.merge_cells("C4:F5")
    c = ws.cell(row=4, column=3, value="Newsreel Relevance Accuracy\n(0 = wrong, 2 = correct)")
    c.font = Font(size=11); c.alignment = CENTER; c.border = THIN

    # Error breakdown
    correct = sum(1 for i in range(1, 11) if NR_SCORES.get(i, ("?", 0, ""))[1] == 2)
    wrong   = sum(1 for i in range(1, 11) if NR_SCORES.get(i, ("?", 0, ""))[1] == 0)
    bline   = sum(1 for i in range(1, 11) if NR_SCORES.get(i, ("?", 0, ""))[1] == 1)

    ws.merge_cells("A7:F7")
    c = ws.cell(row=7, column=1, value="Error Breakdown")
    c.font = Font(bold=True, size=12); c.fill = SUBHD; c.alignment = WRAP_C; c.border = THIN

    for col, (label, val) in enumerate([
        ("Correct", correct), ("Wrong", wrong), ("Borderline", bline)
    ], 1):
        ws.cell(row=8, column=col, value=label).font = BOLD
        ws.cell(row=9, column=col, value=val).font   = Font(size=14, bold=True)
        for row in (8, 9):
            ws.cell(row=row, column=col).border = THIN
            ws.cell(row=row, column=col).alignment = CENTER

    # Error pattern analysis
    ws.merge_cells("A11:F11")
    c = ws.cell(row=11, column=1, value="Error Pattern Analysis")
    c.font = Font(bold=True, size=12); c.fill = SUBHD; c.alignment = WRAP_C; c.border = THIN

    errors = [
        ("False Negative (said No, should Yes)", "Rows 5, 8 — NAIC advocacy letters to Congress",
         "Agent treats government affairs letters as non-newsworthy. These policy positions are relevant to insurance regulation."),
        ("False Positive (said Yes, should No)", "Rows 4, 6, 9 — calendar, routine agenda, editorial corrections",
         "Agent marks scheduling and administrative documents as newsworthy. Needs stricter 'substantive content' filter."),
    ]

    row_out = 12
    for label, rows_str, desc in errors:
        c = ws.cell(row=row_out, column=1, value=label)
        c.font = BOLD; c.alignment = WRAP; c.border = THIN
        ws.merge_cells(start_row=row_out, start_column=1, end_row=row_out, end_column=2)

        c2 = ws.cell(row=row_out, column=3, value=rows_str)
        c2.font = NORM; c2.alignment = WRAP; c2.border = THIN
        ws.merge_cells(start_row=row_out, start_column=3, end_row=row_out, end_column=6)
        ws.row_dimensions[row_out].height = 18
        row_out += 1

        c3 = ws.cell(row=row_out, column=1, value=desc)
        c3.font = Font(size=9, italic=True, color="595959")
        c3.alignment = WRAP; c3.border = THIN
        ws.merge_cells(start_row=row_out, start_column=1, end_row=row_out, end_column=6)
        ws.row_dimensions[row_out].height = 36
        row_out += 1

    # Improvement notes
    ws.merge_cells(f"A{row_out}:F{row_out}")
    c = ws.cell(row=row_out, column=1, value="Recommendations")
    c.font = Font(bold=True, size=12); c.fill = SUBHD; c.alignment = WRAP_C; c.border = THIN
    ws.row_dimensions[row_out].height = 18
    row_out += 1

    recs_text = [
        "1. Add 'NAIC letters to Congress / government affairs advocacy' as an explicit Yes example in the document agent system prompt.",
        "2. Add 'scheduling calendars, routine meeting agendas, editorial corrections' as explicit No examples.",
        "3. The agent's newsreel_relevance details field is blank on all 10 rows — adding a required explanation would make errors easier to diagnose.",
        "4. Most PDFs returned only 35 chars (binary). Fetching PDF text via a text extraction step would let reviewers verify the judgment independently.",
    ]
    for rec in recs_text:
        c = ws.cell(row=row_out, column=1, value=rec)
        c.font = SMALL; c.alignment = WRAP; c.border = THIN
        ws.merge_cells(start_row=row_out, start_column=1, end_row=row_out, end_column=6)
        ws.row_dimensions[row_out].height = 30
        row_out += 1

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 24
    ws.column_dimensions["D"].width = 24
    ws.column_dimensions["E"].width = 24
    ws.column_dimensions["F"].width = 24


def main():
    with open(DATA_PATH) as f:
        records = json.load(f)

    wb = Workbook()
    wb.remove(wb.active)

    pct = sheet_newsreel(wb, records)
    sheet_detail(wb, records)
    sheet_summary(wb, records, pct)

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUTPUT_PATH)
    print(f"Saved → {OUTPUT_PATH}")
    print(f"Newsreel relevance accuracy: {pct}%")

    correct = sum(1 for i in range(1, 11) if NR_SCORES.get(i, ("?", 0, ""))[1] == 2)
    wrong   = sum(1 for i in range(1, 11) if NR_SCORES.get(i, ("?", 0, ""))[1] == 0)
    print(f"Correct: {correct}/10  |  Wrong: {wrong}/10")


if __name__ == "__main__":
    main()
