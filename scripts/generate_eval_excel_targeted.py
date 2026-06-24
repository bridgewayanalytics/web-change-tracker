"""
Generate scored accuracy eval Excel for the three targeted field evals:
  Sheet 1 — Chronicle Topics  (7 rows, agenda-item alerts only)
  Sheet 2 — Event Fields      (10 rows, meeting/agenda alerts only)
  Sheet 3 — Document Fields   (10 rows, alerts with a real document)
  Sheet 4 — Summary           (field accuracy per sheet)

Usage:
    python3 scripts/generate_eval_excel_targeted.py

Reads:
    analysis/accuracy_eval/eval_data_chronicle_topics.json
    analysis/accuracy_eval/eval_data_events.json
    analysis/accuracy_eval/eval_data_documents.json

Writes:
    analysis/accuracy_eval/accuracy_audit_targeted.xlsx
"""

import json
from datetime import datetime
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
BASE = Path("analysis/accuracy_eval")
OUT  = BASE / "accuracy_audit_targeted.xlsx"

# ---------------------------------------------------------------------------
# Colors
# ---------------------------------------------------------------------------
GREEN  = "C6EFCE"
YELLOW = "FFEB9C"
RED    = "FFC7CE"
GRAY   = "D9D9D9"
HEADER = "1F4E79"
SUBHDR = "2E75B6"
WHITE  = "FFFFFF"

def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

thin = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"),  bottom=Side(style="thin"),
)

# ---------------------------------------------------------------------------
# Score rendering helpers
# ---------------------------------------------------------------------------
def score_cell(ws, row, col, score, correction):
    cell = ws.cell(row=row, column=col)
    if score == 2:
        cell.value = "✓ Correct"
        cell.fill = fill(GREEN)
    elif score == 1:
        cell.value = f"△ Partial"
        cell.fill = fill(YELLOW)
        if correction:
            cell.value = f"△ Partial — {correction}"
    elif score == 0:
        cell.value = f"✗ Wrong"
        cell.fill = fill(RED)
        cell.font = Font(bold=True, color="9C0006")
        if correction:
            cell.value = f"✗ {correction}"
    else:  # N/A
        cell.value = "N/A"
        cell.fill = fill(GRAY)
    cell.alignment = Alignment(wrap_text=True, vertical="top")
    cell.border = thin

def hdr(ws, row, col, text, bg=HEADER, fg=WHITE, bold=True, size=10):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(bg)
    c.font = Font(bold=bold, color=fg, size=size)
    c.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    c.border = thin
    return c

def val_cell(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=str(text) if text is not None else "")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    c.border = thin
    return c

# ---------------------------------------------------------------------------
# Human-readable value helpers
# ---------------------------------------------------------------------------
def fmt_dt(val):
    if not val or val == "N/A":
        return str(val or "N/A")
    try:
        dt = datetime.fromisoformat(str(val).replace("Z", "+00:00"))
        return dt.strftime("%B %d, %Y %-I:%M %p ET").replace("ET", "ET")
    except Exception:
        return str(val)

def fmt_title(val):
    if isinstance(val, dict):
        t = val.get("title", "")
        s = val.get("status", "")
        return f"{t} [{s}]" if s else str(t)
    return str(val or "")

def fmt_topics(items):
    if not isinstance(items, list):
        return str(items or "N/A")
    lines = []
    for item in items:
        if not isinstance(item, dict):
            continue
        status = item.get("status", "")
        title  = item.get("agenda_item_title", "")
        topics = item.get("chronicle_topics") or []
        topic_str = ", ".join(t for t in topics if t) if topics else "N/A"
        lines.append(f"[{status}] {title}\n  → {topic_str}")
    return "\n".join(lines) if lines else "N/A"

def fmt_url(val):
    s = str(val or "N/A")
    return s[:120] + "…" if len(s) > 120 else s

# ---------------------------------------------------------------------------
# Scores
# ---------------------------------------------------------------------------

# Chronicle Topics: (row_index) → (score, correction)
# Score covers the entire row (all agenda items together)
CT_SCORES = {
    1: (0,  "N/A for NCRR open meeting — this is a NAIC Climate Initiative; "
            "should tag 'NAIC Climate Initiatives'"),
    2: (2,  ""),   # AG 55 has no dedicated topic; N/A correct
    3: (0,  "Presentation from S&P on Private Credit — should tag 'Standard & Poor (S&P)'"),
    4: (1,  "CLO topic correct; 'RBC Covariance & Asset Concentration Risk' does not "
            "match taxonomy exactly — should be 'Life RBC Covariance & Asset Concentration Risk'"),
    5: (1,  "VM-22 items (APF 2025-18, 2025-20) correctly tagged; APF 2025-14 (VA Scope) "
            "tagged as 'RBC C-3' is questionable — VA scope clarification is primarily a VM "
            "issue, not C-3 interest-rate risk; Group Annuity memo tagged VM-22 "
            "when group annuities fall outside VM-22 scope"),
    6: (2,  ""),   # 'Negative IMR' + 'Credit for Reinsurance' both correct
    7: (0,  "'Climate impact disclosures discussion' and 'Severe convective storm impact "
            "analysis update' both warrant 'NAIC Climate Initiatives'; all N/A missed"),
}

# Events: (row_index, field) → (score, correction)
EV_SCORES = {
    # event_title
    (1,  "event_title"): (1,  "Descriptive and accurate but no standard format — "
                               "joint-meeting title is very long (84 chars); no naming convention applied"),
    (2,  "event_title"): (0,  "'NAIC Interim Meeting:' prefix is agent-invented — "
                               "NAIC page shows the group name without this prefix"),
    (3,  "event_title"): (0,  "'NAIC LATF | June 11, 2026 Public Webex Meeting' — "
                               "pipe format with embedded date is non-standard"),
    (4,  "event_title"): (0,  "'Life Actuarial (A) Task Force - 06252026' — "
                               "date-code suffix is agent-invented, not on the NAIC page"),
    (5,  "event_title"): (0,  "'Financial Stability (E) Task Force - 06152026' — "
                               "date-code suffix is agent-invented, not on the NAIC page"),
    (6,  "event_title"): (2,  ""),  # 'Reinsurance (E) Task Force Meeting' matches page text
    (7,  "event_title"): (0,  "'NAIC Interim Meeting: Reinsurance (E) Task Force Meeting' — "
                               "same page, same event as Row 6 but agent added the prefix this time; "
                               "format inconsistency across alerts for the same meeting"),
    (8,  "event_title"): (0,  "'NAIC LATF | June 11, 2026 Public Webex Meeting' — "
                               "same non-standard pipe format as Row 3"),
    (9,  "event_title"): (0,  "'NAIC Interim Meeting: Life Actuarial (A) Task Force - 06182026' — "
                               "double format error: prefix AND date-code suffix combined"),
    (10, "event_title"): (0,  "'NAIC LATF | ' — incomplete title, missing meeting description "
                               "after the pipe separator"),
    # event_start_date_time (all verified correct or plausible from page)
    (1,  "event_start_date_time"):  (2, ""),
    (2,  "event_start_date_time"):  (2, ""),  # page confirms 3:00 PM ET July 15
    (3,  "event_start_date_time"):  (2, ""),
    (4,  "event_start_date_time"):  (2, ""),
    (5,  "event_start_date_time"):  (2, ""),
    (6,  "event_start_date_time"):  (2, ""),  # page confirms 12:00 PM ET June 22
    (7,  "event_start_date_time"):  (2, ""),
    (8,  "event_start_date_time"):  (2, ""),
    (9,  "event_start_date_time"):  (2, ""),
    (10, "event_start_date_time"):  (2, ""),
    # event_end_date_time
    (1,  "event_end_date_time"):    (2, ""),
    (2,  "event_end_date_time"):    (2, ""),
    (3,  "event_end_date_time"):    (2, ""),  # 1-hour duration matches page "Expected Duration: 1 hour"
    (4,  "event_end_date_time"):    (2, ""),
    (5,  "event_end_date_time"):    (2, ""),
    (6,  "event_end_date_time"):    (2, ""),
    (7,  "event_end_date_time"):    (2, ""),
    (8,  "event_end_date_time"):    (2, ""),
    (9,  "event_end_date_time"):    (2, ""),  # 1.5-hour correctly captured after extension
    (10, "event_end_date_time"):    (2, ""),
    # event_url (Webex link — all correctly extracted)
    (1,  "event_url"):  (2, ""),
    (2,  "event_url"):  (2, ""),
    (3,  "event_url"):  (2, ""),
    (4,  "event_url"):  (2, ""),
    (5,  "event_url"):  (2, ""),
    (6,  "event_url"):  (2, ""),
    (7,  "event_url"):  (2, ""),
    (8,  "event_url"):  (2, ""),
    (9,  "event_url"):  (2, ""),
    (10, "event_url"):  (2, ""),
    # event_call_in_number_access_code (N/A for all modern Webex — correct)
    (1,  "event_call_in_number_access_code"): (2, ""),
    (2,  "event_call_in_number_access_code"): (2, ""),
    (3,  "event_call_in_number_access_code"): (2, ""),
    (4,  "event_call_in_number_access_code"): (2, ""),
    (5,  "event_call_in_number_access_code"): (2, ""),
    (6,  "event_call_in_number_access_code"): (2, ""),
    (7,  "event_call_in_number_access_code"): (2, ""),
    (8,  "event_call_in_number_access_code"): (2, ""),
    (9,  "event_call_in_number_access_code"): (2, ""),
    (10, "event_call_in_number_access_code"): (2, ""),
}

# Documents: (row_index, field) → (score, correction)
DOC_SCORES = {
    # library_item_preliminary_title
    (1,  "library_item_preliminary_title"): (2, ""),
    (2,  "library_item_preliminary_title"): (2, ""),
    (3,  "library_item_preliminary_title"): (2, ""),
    (4,  "library_item_preliminary_title"): (0,  "Title is the raw filename "
                                                   "('GOES Model Change Templates 052826_0.xlsx') — "
                                                   "should be a human-readable description "
                                                   "e.g. 'GOES 2026 Model Change Exposure Draft'"),
    (5,  "library_item_preliminary_title"): (2, ""),   # 'Old: ...' is correct for removed entry
    (6,  "library_item_preliminary_title"): (2, ""),
    (7,  "library_item_preliminary_title"): (2, ""),
    (8,  "library_item_preliminary_title"): (2, ""),   # APF abbreviations acceptable in context
    (9,  "library_item_preliminary_title"): (0,  "Title is 'Updated RBCIREWG 04-10-26 Agenda&Materials' — "
                                                   "uses internal acronym 'RBCIREWG'; should read "
                                                   "'Risk-Based Capital Investment Risk and Evaluation (E) "
                                                   "Working Group Agenda & Materials April 10, 2026'"),
    (10, "library_item_preliminary_title"): (1,  "'Proposal 2026-12-IRE CLO Factor' is partially descriptive "
                                                   "but 'IRE' is an internal abbreviation; "
                                                   "should spell out 'Investment Risk Evaluation'"),
    # library_item_url (all correct)
    (1,  "library_item_url"): (2, ""),
    (2,  "library_item_url"): (2, ""),
    (3,  "library_item_url"): (2, ""),
    (4,  "library_item_url"): (2, ""),
    (5,  "library_item_url"): (2, ""),
    (6,  "library_item_url"): (2, ""),
    (7,  "library_item_url"): (2, ""),
    (8,  "library_item_url"): (2, ""),
    (9,  "library_item_url"): (2, ""),
    (10, "library_item_url"): (2, ""),
    # library_items_file_name (all correct — agent extracts verbatim from page)
    (1,  "library_items_file_name"): (2, ""),
    (2,  "library_items_file_name"): (2, ""),
    (3,  "library_items_file_name"): (2, ""),
    (4,  "library_items_file_name"): (2, ""),
    (5,  "library_items_file_name"): (2, ""),
    (6,  "library_items_file_name"): (2, ""),
    (7,  "library_items_file_name"): (2, ""),
    (8,  "library_items_file_name"): (2, ""),
    (9,  "library_items_file_name"): (2, ""),
    (10, "library_items_file_name"): (2, ""),
}

# ---------------------------------------------------------------------------
# Sheet builders
# ---------------------------------------------------------------------------

def build_chronicle_sheet(ws, data):
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 70
    ws.row_dimensions[1].height = 22

    # Title banner
    ws.merge_cells("A1:C1")
    c = ws["A1"]
    c.value = "TARGETED EVAL — Chronicle Topics  |  7 rows from agenda-type alerts"
    c.fill = fill(HEADER)
    c.font = Font(bold=True, color=WHITE, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")

    # Column headers
    for col, label in enumerate(["Row / Type", "Run ID", "Agenda Items & Topics Assigned"], 1):
        hdr(ws, 2, col, label, bg=SUBHDR)

    row_out = 3
    scores_for_summary = []

    for rec in data:
        idx  = rec["index"]
        a    = rec["alert"]
        atype = str(a.get("alert_type") or "")
        run_id = str(a.get("run_id") or "")[-8:]
        dt   = str(a.get("alert_date_time") or "")[:10]

        score, correction = CT_SCORES.get(idx, (2, ""))
        scores_for_summary.append(score)

        # Header row
        ws.merge_cells(f"A{row_out}:C{row_out}")
        c = ws.cell(row=row_out, column=1,
                    value=f"ROW {idx}  ·  {atype}  ·  {dt}")
        c.fill = fill("D6E4F7")
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(vertical="center")
        c.border = thin
        row_out += 1

        # Agent said row
        topics_text = fmt_topics(a.get("agenda_item_title_chronicle_topics"))
        val_cell(ws, row_out, 1, "Agent Said")
        ws.cell(row=row_out, column=1).fill = fill("F2F2F2")
        val_cell(ws, row_out, 2, run_id)
        val_cell(ws, row_out, 3, topics_text)
        ws.row_dimensions[row_out].height = max(60, topics_text.count("\n") * 14 + 20)
        row_out += 1

        # Score row
        ws.cell(row=row_out, column=1, value="Score").fill = fill("F2F2F2")
        ws.cell(row=row_out, column=1).border = thin
        ws.cell(row=row_out, column=1).alignment = Alignment(vertical="top")
        val_cell(ws, row_out, 2, "")
        score_cell(ws, row_out, 3, score, correction)
        ws.row_dimensions[row_out].height = 50 if correction else 20
        row_out += 1

    # Summary at bottom
    row_out += 1
    ws.cell(row=row_out, column=1, value="Field Accuracy").font = Font(bold=True)
    pts = sum(scores_for_summary)
    max_pts = len(scores_for_summary) * 2
    pct = pts / max_pts * 100 if max_pts else 0
    ws.cell(row=row_out, column=3,
            value=f"Chronicle Topics: {pct:.0f}%  "
                  f"({sum(s==2 for s in scores_for_summary)} correct, "
                  f"{sum(s==1 for s in scores_for_summary)} partial, "
                  f"{sum(s==0 for s in scores_for_summary)} wrong  of {len(scores_for_summary)} rows)")

    return scores_for_summary


def build_events_sheet(ws, data):
    FIELDS = [
        ("event_title",                    "Event Title (format)"),
        ("event_start_date_time",          "Start Time"),
        ("event_end_date_time",            "End Time"),
        ("event_url",                      "Webex URL"),
        ("event_call_in_number_access_code", "Call-In / Access Code"),
    ]

    # Column widths
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    for i in range(len(FIELDS)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 32

    ws.row_dimensions[1].height = 22
    ws.merge_cells(f"A1:{get_column_letter(2 + len(FIELDS))}1")
    c = ws["A1"]
    c.value = "TARGETED EVAL — Event Fields  |  10 rows from meeting/agenda alerts"
    c.fill = fill(HEADER)
    c.font = Font(bold=True, color=WHITE, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")

    for col, (_, label) in enumerate(FIELDS, 3):
        hdr(ws, 2, col, label, bg=SUBHDR)
    hdr(ws, 2, 1, "Row / Type", bg=SUBHDR)
    hdr(ws, 2, 2, "Run ID",     bg=SUBHDR)

    row_out = 3
    field_scores = {f: [] for f, _ in FIELDS}

    for rec in data:
        idx   = rec["index"]
        a     = rec["alert"]
        atype = str(a.get("alert_type") or "")
        run_id = str(a.get("run_id") or "")[-8:]
        dt    = str(a.get("alert_date_time") or "")[:10]

        # Header row
        ws.merge_cells(f"A{row_out}:{get_column_letter(2 + len(FIELDS))}{row_out}")
        c = ws.cell(row=row_out, column=1,
                    value=f"ROW {idx}  ·  {atype}  ·  {dt}")
        c.fill = fill("D6E4F7")
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(vertical="center")
        c.border = thin
        row_out += 1

        # Agent said row
        val_cell(ws, row_out, 1, "Agent Said").fill = fill("F2F2F2")
        ws.cell(row=row_out, column=1).fill = fill("F2F2F2")
        val_cell(ws, row_out, 2, run_id)
        for col, (fkey, _) in enumerate(FIELDS, 3):
            raw = a.get(fkey)
            if fkey == "event_start_date_time" or fkey == "event_end_date_time":
                display = fmt_dt(raw)
            elif fkey == "event_url":
                display = fmt_url(raw)
            else:
                display = str(raw or "N/A")
            val_cell(ws, row_out, col, display)
        ws.row_dimensions[row_out].height = 32
        row_out += 1

        # Score row
        ws.cell(row=row_out, column=1, value="Score").fill = fill("F2F2F2")
        ws.cell(row=row_out, column=1).border = thin
        ws.cell(row=row_out, column=1).alignment = Alignment(vertical="top")
        val_cell(ws, row_out, 2, "")
        for col, (fkey, _) in enumerate(FIELDS, 3):
            s, corr = EV_SCORES.get((idx, fkey), (2, ""))
            score_cell(ws, row_out, col, s, corr)
            field_scores[fkey].append(s)
        ws.row_dimensions[row_out].height = 50
        row_out += 1

    # Summary at bottom
    row_out += 1
    ws.cell(row=row_out, column=1, value="Field Accuracy").font = Font(bold=True)
    for col, (fkey, label) in enumerate(FIELDS, 3):
        scores = field_scores[fkey]
        pts = sum(scores)
        max_pts = len(scores) * 2
        pct = pts / max_pts * 100 if max_pts else 0
        c = ws.cell(row=row_out, column=col, value=f"{label}: {pct:.0f}%")
        c.font = Font(bold=True)
        if pct >= 95:    c.fill = fill(GREEN)
        elif pct >= 80:  c.fill = fill("92D050")
        elif pct >= 60:  c.fill = fill(YELLOW)
        else:            c.fill = fill(RED)

    return field_scores


def build_documents_sheet(ws, data):
    FIELDS = [
        ("library_item_preliminary_title", "Document Title"),
        ("library_item_url",               "Document URL"),
        ("library_items_file_name",        "Filename"),
    ]

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 24
    for i in range(len(FIELDS)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 44

    ws.row_dimensions[1].height = 22
    ws.merge_cells(f"A1:{get_column_letter(2 + len(FIELDS))}1")
    c = ws["A1"]
    c.value = "TARGETED EVAL — Document Fields  |  10 rows from alerts with a document"
    c.fill = fill(HEADER)
    c.font = Font(bold=True, color=WHITE, size=12)
    c.alignment = Alignment(horizontal="center", vertical="center")

    for col, (_, label) in enumerate(FIELDS, 3):
        hdr(ws, 2, col, label, bg=SUBHDR)
    hdr(ws, 2, 1, "Row / Type", bg=SUBHDR)
    hdr(ws, 2, 2, "Run ID",     bg=SUBHDR)

    row_out = 3
    field_scores = {f: [] for f, _ in FIELDS}

    for rec in data:
        idx   = rec["index"]
        a     = rec["alert"]
        atype = str(a.get("alert_type") or "")
        run_id = str(a.get("run_id") or "")[-8:]
        dt    = str(a.get("alert_date_time") or "")[:10]

        ws.merge_cells(f"A{row_out}:{get_column_letter(2 + len(FIELDS))}{row_out}")
        c = ws.cell(row=row_out, column=1,
                    value=f"ROW {idx}  ·  {atype}  ·  {dt}")
        c.fill = fill("D6E4F7")
        c.font = Font(bold=True, size=10)
        c.alignment = Alignment(vertical="center")
        c.border = thin
        row_out += 1

        # Agent said row
        ws.cell(row=row_out, column=1, value="Agent Said").fill = fill("F2F2F2")
        ws.cell(row=row_out, column=1).fill = fill("F2F2F2")
        ws.cell(row=row_out, column=1).border = thin
        ws.cell(row=row_out, column=1).alignment = Alignment(vertical="top")
        val_cell(ws, row_out, 2, run_id)
        for col, (fkey, _) in enumerate(FIELDS, 3):
            raw = a.get(fkey)
            if fkey == "library_item_preliminary_title":
                display = fmt_title(raw)
            elif fkey == "library_item_url":
                display = fmt_url(raw)
            else:
                display = str(raw or "N/A")
            val_cell(ws, row_out, col, display)
        ws.row_dimensions[row_out].height = 32
        row_out += 1

        # Score row
        ws.cell(row=row_out, column=1, value="Score").fill = fill("F2F2F2")
        ws.cell(row=row_out, column=1).border = thin
        ws.cell(row=row_out, column=1).alignment = Alignment(vertical="top")
        val_cell(ws, row_out, 2, "")
        for col, (fkey, _) in enumerate(FIELDS, 3):
            s, corr = DOC_SCORES.get((idx, fkey), (2, ""))
            score_cell(ws, row_out, col, s, corr)
            field_scores[fkey].append(s)
        ws.row_dimensions[row_out].height = 50
        row_out += 1

    row_out += 1
    ws.cell(row=row_out, column=1, value="Field Accuracy").font = Font(bold=True)
    for col, (fkey, label) in enumerate(FIELDS, 3):
        scores = field_scores[fkey]
        pts = sum(scores)
        max_pts = len(scores) * 2
        pct = pts / max_pts * 100 if max_pts else 0
        c = ws.cell(row=row_out, column=col, value=f"{label}: {pct:.0f}%")
        c.font = Font(bold=True)
        if pct >= 95:    c.fill = fill(GREEN)
        elif pct >= 80:  c.fill = fill("92D050")
        elif pct >= 60:  c.fill = fill(YELLOW)
        else:            c.fill = fill(RED)

    return field_scores


def build_summary_sheet(ws, ct_scores, ev_field_scores, doc_field_scores):
    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14

    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = "TARGETED EVAL — Field Accuracy Summary"
    c.fill = fill(HEADER)
    c.font = Font(bold=True, color=WHITE, size=13)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 24

    for col, label in enumerate(["Field", "% Correct", "Correct", "Partial", "Wrong"], 1):
        hdr(ws, 2, col, label, bg=SUBHDR)

    rows = [
        ("── CHRONICLE TOPICS ──", None),
        ("Agenda Items & Chronicle Topics", ct_scores),
        ("── EVENT FIELDS ──", None),
        ("Event Title (format)", ev_field_scores.get("event_title")),
        ("Start Time",           ev_field_scores.get("event_start_date_time")),
        ("End Time",             ev_field_scores.get("event_end_date_time")),
        ("Webex URL",            ev_field_scores.get("event_url")),
        ("Call-In / Access Code",ev_field_scores.get("event_call_in_number_access_code")),
        ("── DOCUMENT FIELDS ──", None),
        ("Document Title",       doc_field_scores.get("library_item_preliminary_title")),
        ("Document URL",         doc_field_scores.get("library_item_url")),
        ("Filename",             doc_field_scores.get("library_items_file_name")),
    ]

    out_row = 3
    for label, scores in rows:
        if scores is None:
            # section header
            ws.merge_cells(f"A{out_row}:E{out_row}")
            c = ws.cell(row=out_row, column=1, value=label)
            c.fill = fill("BDD7EE")
            c.font = Font(bold=True, size=10)
            c.alignment = Alignment(vertical="center")
            c.border = thin
            out_row += 1
            continue

        pts = sum(scores)
        max_pts = len(scores) * 2
        pct = pts / max_pts * 100 if max_pts else 0
        n_correct  = sum(s == 2 for s in scores)
        n_partial  = sum(s == 1 for s in scores)
        n_wrong    = sum(s == 0 for s in scores)

        ws.cell(row=out_row, column=1, value=label).border = thin
        ws.cell(row=out_row, column=1).alignment = Alignment(vertical="center")

        pct_cell = ws.cell(row=out_row, column=2, value=f"{pct:.0f}%")
        pct_cell.font = Font(bold=True)
        pct_cell.border = thin
        pct_cell.alignment = Alignment(horizontal="center", vertical="center")
        if pct >= 95:    pct_cell.fill = fill(GREEN)
        elif pct >= 80:  pct_cell.fill = fill("92D050")
        elif pct >= 60:  pct_cell.fill = fill(YELLOW)
        else:            pct_cell.fill = fill(RED)

        for col, val in [(3, n_correct), (4, n_partial), (5, n_wrong)]:
            c = ws.cell(row=out_row, column=col, value=val)
            c.border = thin
            c.alignment = Alignment(horizontal="center", vertical="center")

        out_row += 1

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main():
    ct_data  = json.loads((BASE / "eval_data_chronicle_topics.json").read_text())
    ev_data  = json.loads((BASE / "eval_data_events.json").read_text())
    doc_data = json.loads((BASE / "eval_data_documents.json").read_text())

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_summary = wb.create_sheet("Summary")
    ws_ct      = wb.create_sheet("Chronicle Topics")
    ws_ev      = wb.create_sheet("Event Fields")
    ws_doc     = wb.create_sheet("Document Fields")

    ct_scores      = build_chronicle_sheet(ws_ct, ct_data)
    ev_field_scores = build_events_sheet(ws_ev, ev_data)
    doc_field_scores = build_documents_sheet(ws_doc, doc_data)
    build_summary_sheet(ws_summary, ct_scores, ev_field_scores, doc_field_scores)

    # Reorder so Summary is first
    wb._sheets = [ws_summary, ws_ct, ws_ev, ws_doc]

    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Saved → {OUT}")

    # Print quick summary
    print("\nField Accuracy:")
    ct_pts = sum(ct_scores)
    print(f"  Chronicle Topics:  {ct_pts / (len(ct_scores)*2) * 100:.0f}%  "
          f"({sum(s==2 for s in ct_scores)} correct, "
          f"{sum(s==1 for s in ct_scores)} partial, "
          f"{sum(s==0 for s in ct_scores)} wrong)")
    for fkey, label in [
        ("event_title",                    "Event Title (format)"),
        ("event_start_date_time",          "Start Time"),
        ("event_end_date_time",            "End Time"),
        ("event_url",                      "Webex URL"),
        ("event_call_in_number_access_code", "Call-In"),
    ]:
        s = ev_field_scores[fkey]
        print(f"  {label:30s}  {sum(s)/len(s)/2*100:.0f}%")
    for fkey, label in [
        ("library_item_preliminary_title", "Document Title"),
        ("library_item_url",               "Document URL"),
        ("library_items_file_name",        "Filename"),
    ]:
        s = doc_field_scores[fkey]
        print(f"  {label:30s}  {sum(s)/len(s)/2*100:.0f}%")


if __name__ == "__main__":
    main()
