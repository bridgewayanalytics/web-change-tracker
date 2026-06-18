"""
Store pipeline alert output to S3 in a UI-ready structure.

Structure (all under CHANGELOG_BUCKET, falling back to BUBBLE_ARTIFACT_BUCKET):

  runs/YYYY/MM/DD/<run_id>/alerts.json
      Array of alert objects — one per changed page that produced agent output.
      This is the primary file for the Alerts Table dashboard UI.

  alerts/alerts_table.jsonl
      Single growing flat JSONL — every run appends its rows here.
      One line per alert row using the exact field names from the Bubble field
      inventory Alerts table. One row per library item detected (or one row per
      alert when no library items). This is the primary stakeholder-facing file.

  pages/<target_id>/YYYY/MM/DD/<run_id>/agent_output.json
      Full web tracking agent output for one page.

  pages/<target_id>/YYYY/MM/DD/<run_id>/doc_extractions.json
      Document agent results for library items found on that page.

Never raises — all failures are logged as warnings.
"""

import json
import logging
import os
from datetime import datetime, timezone

log = logging.getLogger(__name__)


_DEFAULT_BUCKET = "web-change-tracker-prod-artifacts-815039343351"


def _get_bucket() -> str:
    return (
        os.environ.get("CHANGELOG_BUCKET", "").strip()
        or os.environ.get("BUBBLE_ARTIFACT_BUCKET", "").strip()
        or _DEFAULT_BUCKET
    )


def _s3_client():
    import boto3
    return boto3.client("s3", region_name=os.environ.get("AWS_REGION", "us-east-1"))


def _date_prefix(run_timestamp: int) -> str:
    dt = datetime.fromtimestamp(run_timestamp, tz=timezone.utc)
    return f"{dt.year:04d}/{dt.month:02d}/{dt.day:02d}"


def _put(client, bucket: str, key: str, body: bytes, content_type: str, run_id: str) -> None:
    client.put_object(
        Bucket=bucket,
        Key=key,
        Body=body,
        ContentType=content_type,
        Metadata={"run_id": run_id},
    )


def _put_json(client, bucket: str, key: str, obj: object, run_id: str) -> None:
    _put(client, bucket, key,
         json.dumps(obj, indent=2, ensure_ascii=False).encode("utf-8"),
         "application/json", run_id)


def _flatten_val(val) -> str | bool | None:
    """
    Coerce agent output values to scalar types safe for JSONL / Excel cells.

    Preserves the agent's actual output — null stays null, empty string stays
    empty string — so the dashboard shows exactly what the agent said.
    Lists are joined with ", " for display (or kept as "" if empty).
    """
    if isinstance(val, bool):
        return val
    if isinstance(val, list):
        items = [str(x) for x in val if x is not None and x != ""]
        return ", ".join(items) if items else ""
    return val


def _build_rows_for_single_alert(
    agent_output: dict,
    run_id: str,
    run_timestamp_iso: str,
    target_id: str,
    source_url: str,
    config_hash: str = "",
) -> list[dict]:
    """
    Build alert_table rows from a single alert dict.

    Handles two schema formats:
    - **Flat schema (May 2026+):** All fields are top-level (event_title,
      library_item_url, etc.). Stored verbatim — one row per alert.
    - **Nested schema (pre-May 2026):** events[], library_items[], agenda_items[]
      arrays. Exploded: one row per library item, first-item fields flattened
      with event_*/agenda_item_* prefixes.
    """
    _NESTED = {"events", "library_items", "agenda_items"}

    # Detect schema format: if any nested array key is present with actual
    # list content, use the old nested-array path.
    has_nested = any(
        isinstance(agent_output.get(k), list) and agent_output.get(k)
        for k in _NESTED
    )

    base: dict = {
        "run_id": run_id,
        "run_timestamp": run_timestamp_iso,
        "target_id": target_id,
        "source_url": source_url,
        "config_hash": config_hash,
    }

    if not has_nested:
        # Flat schema — all agent output fields stored verbatim as top-level keys
        for key, val in agent_output.items():
            base[key] = val
        return [base]

    # --- Nested schema (backward compat for old rows) ---

    # All top-level agent output fields (excluding nested arrays)
    for key, val in agent_output.items():
        if key not in _NESTED:
            base[key] = val

    # Full arrays — stored as-is for rerun diff view
    events = agent_output.get("events") or []
    agenda_items = agent_output.get("agenda_items") or []
    base["events"] = events
    base["agenda_items"] = agenda_items

    # First-item flattening for backward compatibility
    if events:
        for key, val in events[0].items():
            base[f"event_{key}"] = _flatten_val(val)
    if agenda_items:
        for key, val in agenda_items[0].items():
            base[f"agenda_item_{key}"] = _flatten_val(val)

    library_items = agent_output.get("library_items") or []
    if not library_items:
        return [base]

    rows: list[dict] = []
    for item in library_items:
        row = dict(base)
        for key, val in item.items():
            row[f"library_item_{key}"] = val
        rows.append(row)

    return rows


def _build_table_rows(
    agent_outputs: list[dict] | dict,
    doc_extractions: list[dict],
    run_id: str,
    run_timestamp_iso: str,
    target_id: str,
    source_url: str,
    config_hash: str = "",
) -> list[dict]:
    """
    Build alert_table rows from web-tracking-agent output.

    Accepts either a list of alert dicts (multi-alert) or a single dict
    (backward compatibility). Delegates to _build_rows_for_single_alert
    for each alert.
    """
    # Normalize to list
    if isinstance(agent_outputs, dict):
        alerts = [agent_outputs]
    else:
        alerts = agent_outputs

    rows: list[dict] = []
    for alert in alerts:
        rows.extend(_build_rows_for_single_alert(
            alert, run_id, run_timestamp_iso, target_id, source_url,
            config_hash=config_hash,
        ))
    return rows


def _build_doc_extraction_rows(
    doc_extractions: list[dict],
    run_id: str,
    run_timestamp_iso: str,
    target_id: str,
    source_url: str,
    agent_call_id: str = "",
) -> list[dict]:
    """
    Build document_extractions_table rows — one per library item processed.

    All fields returned by the document-data-extraction agent are stored verbatim.
    Pipeline metadata (run_id, target_id, etc.) added as anchoring keys.
    """
    rows: list[dict] = []
    for entry in doc_extractions:
        item = entry.get("item") or {}
        extraction = entry.get("extraction") or {}
        if not extraction:
            continue

        row: dict = {
            "run_id": run_id,
            "run_timestamp": run_timestamp_iso,
            "target_id": target_id,
            "source_url": source_url,
            "agent_call_id": agent_call_id,
            "library_item_title": item.get("preliminary_title") or item.get("title") or "",
            "library_item_url": item.get("url") or "",
            "library_item_file_name": item.get("file_name") or "",
        }
        # All document agent output fields verbatim
        for key, val in extraction.items():
            row[key] = val

        rows.append(row)

    return rows


def patch_jsonl_row(
    jsonl_key: str,
    match_fields: dict,
    update_fields: dict,
    bucket: str | None = None,
) -> int:
    """
    Patch rows in a JSONL file in S3 where all match_fields match.

    Downloads the file, updates matching rows in-place, re-uploads.
    Returns the number of rows patched. Never raises — logs warnings on failure.

    Example:
        patch_jsonl_row(
            "alerts/alerts_table.jsonl",
            {"agent_call_id": "abc123"},
            {"ingest_status": "approved"},
        )
    """
    if bucket is None:
        bucket = _get_bucket()
    if not bucket:
        log.warning("patch_jsonl_row: no bucket configured")
        return 0

    try:
        client = _s3_client()
        try:
            body = client.get_object(Bucket=bucket, Key=jsonl_key)["Body"].read().decode("utf-8")
        except client.exceptions.NoSuchKey:
            log.warning("patch_jsonl_row: %s not found in bucket %s", jsonl_key, bucket)
            return 0

        patched = 0
        out_lines: list[str] = []
        for line in body.splitlines():
            line = line.strip()
            if not line:
                continue
            try:
                row = json.loads(line)
            except Exception:
                out_lines.append(line)
                continue

            if all(row.get(k) == v for k, v in match_fields.items()):
                row.update(update_fields)
                patched += 1

            out_lines.append(json.dumps(row, ensure_ascii=False))

        if patched == 0:
            log.warning("patch_jsonl_row: no rows matched %s in %s", match_fields, jsonl_key)
            return 0

        combined = "\n".join(out_lines).encode("utf-8")
        client.put_object(
            Bucket=bucket,
            Key=jsonl_key,
            Body=combined,
            ContentType="application/x-ndjson",
        )
        log.info("patch_jsonl_row: patched %d row(s) in s3://%s/%s", patched, bucket, jsonl_key)
        return patched

    except Exception as exc:
        log.warning("patch_jsonl_row: failed to patch %s: %s", jsonl_key, exc)
        return 0


def store_run_alerts(
    change_events: list[dict],
    run_id: str,
    run_timestamp: int,
) -> str | None:
    """
    Write:
      runs/<date>/<run_id>/alerts.json       — structured alert objects (one per page)
      alerts/alerts_table.jsonl              — growing flat JSONL, appended each run
      pages/<target_id>/<date>/<run_id>/agent_output.json
      pages/<target_id>/<date>/<run_id>/doc_extractions.json

    Only events with __agent_output and alert_type != "No Meaningful Change" are included.
    Never raises — all failures are logged as warnings.

    Returns the S3 URI of alerts.json on success, None if skipped or failed.
    """
    bucket = _get_bucket()
    if not bucket:
        return None

    date_prefix = _date_prefix(run_timestamp)
    run_timestamp_iso = datetime.fromtimestamp(run_timestamp, tz=timezone.utc).isoformat()

    # Compute config hash once per run (empty string if agent not importable)
    config_hash = ""
    try:
        from bubble.page_change_agent import get_config_hash
        config_hash = get_config_hash()
    except Exception:
        pass

    alert_rows: list[dict] = []
    table_rows: list[dict] = []
    doc_table_rows: list[dict] = []

    try:
        client = _s3_client()
    except Exception as e:
        log.warning("alert_s3: could not create S3 client: %s", e)
        return None

    for ev in change_events:
        if "error" in ev:
            continue
        # __agent_output is now a list of alert dicts (or a single dict for backward compat)
        raw_output = ev.get("__agent_output")
        if not raw_output:
            continue
        if isinstance(raw_output, dict):
            agent_alerts = [raw_output]
        else:
            agent_alerts = raw_output

        # Filter out "No Meaningful Change" alerts (handles both alert_type and Alert Type1 keys)
        from bubble.page_change_agent import _is_no_meaningful_change
        agent_alerts = [a for a in agent_alerts if not _is_no_meaningful_change(a)]
        if not agent_alerts:
            continue

        target_id = ev.get("target_id") or "unknown"
        source_url = ev.get("url") or ""
        agent_call_id = agent_alerts[0].get("agent_call_id", "")
        page_key_prefix = f"pages/{target_id}/{date_prefix}/{run_id}"

        # Write per-page agent_output.json (full list of alerts)
        agent_key = f"{page_key_prefix}/agent_output.json"
        try:
            _put_json(client, bucket, agent_key, agent_alerts, run_id)
        except Exception as e:
            log.warning("alert_s3: failed to write agent_output for %s: %s", target_id, e)
            agent_key = None

        # Write per-page doc_extractions.json (if present)
        doc_extractions = ev.get("__doc_extraction") or []
        if doc_extractions:
            doc_key = f"{page_key_prefix}/doc_extractions.json"
            try:
                _put_json(client, bucket, doc_key, doc_extractions, run_id)
            except Exception as e:
                log.warning("alert_s3: failed to write doc_extractions for %s: %s", target_id, e)

        # Structured alert rows (alerts.json) — one per alert
        for agent_output in agent_alerts:
            from bubble.page_change_agent import _get_alert_type
            row: dict = {
                "run_id": run_id,
                "run_timestamp": run_timestamp_iso,
                "target_id": target_id,
                "source_url": source_url,
                "agent_call_id": agent_call_id,
                "alert_type": _get_alert_type(agent_output),
                "alert_title": agent_output.get("alert_title") or agent_output.get("Alert Title") or "",
                "alert_description": agent_output.get("alert_description") or agent_output.get("Alert Description") or "",
                "alert_url": agent_output.get("alert_url") or agent_output.get("Alert URL"),
                "organization": agent_output.get("organization") or agent_output.get("Organization"),
                "alert_date_time": agent_output.get("alert_date_time") or agent_output.get("Alert Date & Time (ET)"),
                "events": agent_output.get("events") or [],
                "library_items": agent_output.get("library_items") or [],
                "agenda_items": agent_output.get("agenda_items") or [],
                "doc_extractions": [e.get("extraction") or {} for e in doc_extractions],
            }
            if agent_key:
                row["detail_s3_key"] = agent_key
            alert_rows.append(row)

        # Alert table rows (alerts_table.jsonl) — one per alert (per library item)
        table_rows.extend(_build_table_rows(
            agent_alerts, doc_extractions,
            run_id, run_timestamp_iso, target_id, source_url,
            config_hash=config_hash,
        ))

        # Document extraction table rows (document_extractions_table.jsonl)
        doc_table_rows.extend(_build_doc_extraction_rows(
            doc_extractions,
            run_id, run_timestamp_iso, target_id, source_url,
            agent_call_id=agent_call_id,
        ))

    # Write alerts.json (per-run structured output)
    alerts_key = f"runs/{date_prefix}/{run_id}/alerts.json"
    try:
        _put_json(client, bucket, alerts_key, alert_rows, run_id)
        uri = f"s3://{bucket}/{alerts_key}"
        log.info("alert_s3: wrote %d alert(s) to %s", len(alert_rows), uri)
    except Exception as e:
        log.warning("alert_s3: failed to write alerts.json: %s", e)
        uri = None

    # Append new rows to the single growing alerts/alerts_table.jsonl
    # S3 has no native append — download existing, append, re-upload.
    # Then regenerate alerts_table.xlsx from the full JSONL.
    if table_rows:
        global_table_key = "alerts/alerts_table.jsonl"
        xlsx_key = "alerts/alerts_table.xlsx"
        try:
            existing_body = b""
            try:
                resp = client.get_object(Bucket=bucket, Key=global_table_key)
                existing_body = resp["Body"].read()
            except client.exceptions.NoSuchKey:
                pass
            except Exception:
                pass

            new_lines = "\n".join(json.dumps(r, ensure_ascii=False) for r in table_rows)
            if existing_body:
                combined = existing_body.rstrip(b"\n") + b"\n" + new_lines.encode("utf-8")
            else:
                combined = new_lines.encode("utf-8")

            _put(client, bucket, global_table_key, combined, "application/x-ndjson", run_id)
            log.info(
                "alert_s3: appended %d row(s) to s3://%s/%s",
                len(table_rows), bucket, global_table_key,
            )

            # Regenerate Excel from the full updated JSONL
            try:
                all_rows = [json.loads(ln) for ln in combined.decode("utf-8").splitlines() if ln.strip()]
                xlsx_bytes = _build_xlsx(all_rows)
                _put(client, bucket, xlsx_key, xlsx_bytes,
                     "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", run_id)
                log.info("alert_s3: updated s3://%s/%s (%d rows)", bucket, xlsx_key, len(all_rows))
            except Exception as e:
                log.warning("alert_s3: failed to write alerts_table.xlsx: %s", e)

        except Exception as e:
            log.warning("alert_s3: failed to update alerts_table.jsonl: %s", e)

    # Append document extraction rows to document_extractions_table.jsonl
    if doc_table_rows:
        doc_table_key = "alerts/document_extractions_table.jsonl"
        try:
            existing_body = b""
            try:
                resp = client.get_object(Bucket=bucket, Key=doc_table_key)
                existing_body = resp["Body"].read()
            except client.exceptions.NoSuchKey:
                pass
            except Exception:
                pass

            new_lines = "\n".join(json.dumps(r, ensure_ascii=False) for r in doc_table_rows)
            if existing_body:
                combined = existing_body.rstrip(b"\n") + b"\n" + new_lines.encode("utf-8")
            else:
                combined = new_lines.encode("utf-8")

            _put(client, bucket, doc_table_key, combined, "application/x-ndjson", run_id)
            log.info(
                "alert_s3: appended %d doc extraction row(s) to s3://%s/%s",
                len(doc_table_rows), bucket, doc_table_key,
            )
        except Exception as e:
            log.warning("alert_s3: failed to update document_extractions_table.jsonl: %s", e)

    return uri


def _build_xlsx(rows: list[dict]) -> bytes:
    """Convert alert table rows to an Excel workbook (bytes).

    Columns are derived dynamically from the row data so that any new fields
    added to the agent output schema appear automatically. Known core columns
    are pinned to the front in a stable order; any extra keys are appended
    alphabetically after them.
    """
    import io
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    if not rows:
        rows = []

    # Core columns in preferred display order (mirrors current flat agent output schema,
    # May 2026+). Any fields not listed here appear after these, sorted alphabetically
    # (this includes old nested-schema aliases for backward compat rows).
    CORE_COLS = [
        "run_id", "run_timestamp", "target_id", "source_url",
        # Top-level alert fields (exact agent output names, flat schema)
        "alert_type", "alert_title", "alert_description", "alert_url",
        "organization", "alert_date_time",
        # event fields
        "event_title", "event_start_date_time", "event_end_date_time",
        "event_timezone", "event_duration", "event_is_full_day",
        "event_url", "event_call_in_number_access_code",
        # agenda item arrays
        "agenda_item_title_chronicle_topics", "agenda_item_title_official",
        "agenda_item_standardized_id", "agenda_item_official_id",
        # library item
        "library_item_preliminary_title", "library_item_url", "library_items_file_name",
        # newsreel relevance
        "is_the_alert_relevant_for_an_art_newsreel_article",
        # doc extraction
        "candidate_chronicles", "candidate_agenda_items",
        # pipeline metadata
        "agent_call_id", "config_hash",
        "recording_s3_key", "transcript_s3_key", "ingest_status",
        "bubble_action", "bubble_sync_status",
    ]

    # Collect all keys present in the data, append any not in CORE_COLS
    all_keys: set[str] = set()
    for row in rows:
        all_keys.update(row.keys())
    core_set = set(CORE_COLS)
    extra_cols = sorted(k for k in all_keys if k not in core_set)
    COLS = CORE_COLS + extra_cols

    # Per-column widths; unknown columns default to 25
    COL_WIDTHS = {
        "run_id": 18, "run_timestamp": 22, "target_id": 22, "source_url": 30,
        "alert_type": 22, "alert_title": 40, "alert_description": 60,
        "alert_url": 30, "organization": 30, "alert_date_time": 20,
        "event_title": 35, "event_start_date_time": 22, "event_end_date_time": 22,
        "event_timezone": 16, "event_duration": 14, "event_is_full_day": 14,
        "event_url": 30, "event_call_in_number_access_code": 28,
        "agenda_item_title_chronicle_topics": 45, "agenda_item_title_official": 40,
        "agenda_item_standardized_id": 22, "agenda_item_official_id": 18,
        "library_item_preliminary_title": 40, "library_item_url": 35,
        "library_items_file_name": 30,
        "is_the_alert_relevant_for_an_art_newsreel_article": 30,
        "candidate_chronicles": 25, "candidate_agenda_items": 25,
        "agent_call_id": 20, "config_hash": 20,
        "recording_s3_key": 35, "transcript_s3_key": 35, "ingest_status": 16,
        "bubble_action": 50, "bubble_sync_status": 16,
    }

    def thin_border():
        s = Side(style="thin", color="D0D0D0")
        return Border(top=s, bottom=s, left=s, right=s)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Alerts"

    # Header row
    for ci, col in enumerate(COLS, 1):
        label = col.replace("_", " ").title()
        c = ws.cell(1, ci, label)
        c.font = Font(name="Calibri", bold=True, size=9, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E40AF")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
    ws.row_dimensions[1].height = 28

    # Data rows
    ALT = "EBF0FF"
    for ri, row in enumerate(rows, 2):
        bg = ALT if ri % 2 == 0 else "FFFFFF"
        for ci, col in enumerate(COLS, 1):
            val = row.get(col, "")
            if isinstance(val, bool):
                val = "Yes" if val else ""
            elif isinstance(val, (list, dict)):
                val = json.dumps(val, ensure_ascii=False)
            c = ws.cell(ri, ci, val)
            c.font = Font(name="Calibri", size=9)
            c.fill = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(vertical="top", wrap_text=True)
            c.border = thin_border()

    # Column widths
    for ci, col in enumerate(COLS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 25)

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
